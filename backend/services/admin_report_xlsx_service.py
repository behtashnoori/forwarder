"""XLSX export helpers for admin report overviews."""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.services import admin_report_overview_service

XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SHEET_SUMMARY = "خلاصه مدیریتی"
SHEET_STATUS = "وضعیت درخواست‌ها"
SHEET_TRANSPORT = "توزیع نوع حمل"
SHEET_LOCATION = "توزیع مبدا و مقصد"
SHEET_EXPERTS = "عملکرد کارشناسان"
SHEET_CUSTOMERS = "مشتریان و درخواست‌ها"
SHEET_TRENDS = "روند زمانی"
SHEET_DETAILS = "جزئیات درخواست‌ها"

MISSING_FIELD_LABEL = admin_report_overview_service.MISSING_FIELD_LABEL


def build_report_xlsx(period: str | None) -> tuple[bytes, str]:
    """Build XLSX bytes from the Phase 12B report overview payload."""
    payload = admin_report_overview_service.get_report_overview_payload(period)
    workbook = Workbook()
    workbook.remove(workbook.active)

    add_summary_sheet(workbook, payload)
    add_table_sheet(
        workbook,
        SHEET_STATUS,
        ["وضعیت", "برچسب فارسی وضعیت", "تعداد درخواست", "درصد از کل", "آخرین تاریخ ثبت در این وضعیت"],
        payload["requests_by_status"],
        ["status", "label_fa", "count", "percent_of_total", "latest_created_at"],
    )
    add_table_sheet(
        workbook,
        SHEET_TRANSPORT,
        ["نوع حمل", "برچسب فارسی نوع حمل", "تعداد درخواست", "درصد از کل", "تعداد تکمیل‌شده", "تعداد در حال بررسی"],
        payload["transport_distribution"],
        ["transport_method", "label_fa", "count", "percent_of_total", "completed_count", "in_progress_count"],
    )
    add_table_sheet(
        workbook,
        SHEET_LOCATION,
        ["نوع موقعیت", "استان", "شهرستان", "شهر", "تعداد درخواست", "درصد از کل", "رایج‌ترین نوع حمل"],
        payload["location_distribution"],
        ["location_type_label_fa", "province", "county", "city", "count", "percent_of_total", "top_transport_method"],
    )
    add_table_sheet(
        workbook,
        SHEET_EXPERTS,
        [
            "شناسه کارشناس",
            "نام کارشناس",
            "نام کاربری",
            "نقش",
            "کل درخواست‌های تخصیص‌یافته",
            "فعال",
            "تکمیل‌شده",
            "ردشده / از دست‌رفته",
            "نرخ تکمیل",
        ],
        payload["expert_performance"],
        [
            "expert_id",
            "expert_name",
            "username",
            "role",
            "assigned_count",
            "active_count",
            "completed_count",
            "lost_or_rejected_count",
            "completion_rate",
        ],
    )
    add_table_sheet(
        workbook,
        SHEET_CUSTOMERS,
        [
            "نام مشتری",
            "شماره تماس",
            "شرکت",
            "تعداد درخواست",
            "آخرین تاریخ درخواست",
            "وضعیت آخرین درخواست",
            "نوع حمل غالب",
            "مبدا غالب",
            "مقصد غالب",
        ],
        payload["customer_concentration"],
        [
            "customer_name",
            "phone",
            "company",
            "request_count",
            "latest_request_date",
            "latest_request_status_label_fa",
            "dominant_transport_method",
            "dominant_origin",
            "dominant_destination",
        ],
    )
    add_table_sheet(
        workbook,
        SHEET_TRENDS,
        ["بازه", "عنوان بازه", "تعداد درخواست"],
        payload["trends"],
        ["bucket", "label", "request_count"],
    )
    add_table_sheet(
        workbook,
        SHEET_DETAILS,
        [
            "شناسه",
            "کد رهگیری",
            "تاریخ ثبت",
            "وضعیت",
            "نام مشتری",
            "شماره تماس",
            "نوع حمل",
            "مبدا استان",
            "مبدا شهرستان",
            "مبدا شهر",
            "مقصد استان",
            "مقصد شهرستان",
            "مقصد شهر",
            "کارشناس",
            "شرح بار",
            "وزن",
            "حجم",
            "ارزش بار",
            "تاریخ بارگیری",
            "تاریخ تحویل",
            "خوانده‌نشده برای کارشناس",
        ],
        payload["request_details"],
        [
            "id",
            "tracking_number",
            "created_at",
            "status_label_fa",
            "customer_name",
            "phone",
            "transport_method",
            "origin_province",
            "origin_county",
            "origin_city",
            "destination_province",
            "destination_county",
            "destination_city",
            "expert_name",
            "cargo_description",
            "weight",
            "volume",
            "cargo_value",
            "pickup_date",
            "delivery_date",
            "has_unread_for_assignee",
        ],
    )

    stream = BytesIO()
    workbook.save(stream)
    filename = f"forwarder-report-{payload['metadata']['period']}.xlsx"
    return stream.getvalue(), filename


def add_summary_sheet(workbook: Workbook, payload: dict[str, Any]) -> None:
    """Add metadata and summary metrics sheet."""
    sheet = workbook.create_sheet(SHEET_SUMMARY)
    configure_sheet(sheet)

    metadata = payload["metadata"]
    summary = payload["summary"]
    rows = [
        ("عنوان گزارش", "گزارش مدیریتی فورواردر"),
        ("دوره", metadata["period_label_fa"]),
        ("شروع بازه", metadata["start_date"]),
        ("پایان بازه", metadata["end_date"]),
        ("زمان تولید", metadata["generated_at"]),
        ("مبنای زمانی", metadata["timezone_basis"]),
        ("", ""),
        ("شاخص", "مقدار"),
        ("کل درخواست‌ها", summary["total_requests"]),
        ("درخواست‌های جدید", summary["new_requests"]),
        ("در انتظار بررسی", summary["pending_review_requests"]),
        ("در حال بررسی", summary["in_progress_requests"]),
        ("تکمیل‌شده", summary["completed_requests"]),
        ("ردشده / از دست‌رفته", summary["lost_or_rejected_requests"]),
        ("بدون کارشناس", summary["unassigned_requests"]),
        ("درخواست‌های ۲۴ ساعت اخیر", summary["requests_last_24h"]),
        ("درخواست‌های ۷ روز اخیر", summary["requests_last_7d"]),
    ]

    for row in rows:
        sheet.append(row)

    sheet["A8"].font = Font(bold=True)
    sheet["B8"].font = Font(bold=True)
    sheet.freeze_panes = "A9"
    auto_size_columns(sheet)


def add_table_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[dict[str, Any]],
    fields: list[str],
) -> None:
    """Add a standard tabular sheet with headers and rows."""
    sheet = workbook.create_sheet(title)
    configure_sheet(sheet)
    sheet.append(headers)
    for row in rows:
        sheet.append([cell_value(row.get(field)) for field in fields])
    style_header_row(sheet)
    sheet.freeze_panes = "A2"
    auto_size_columns(sheet)


def configure_sheet(sheet) -> None:
    """Apply minimal readable sheet configuration."""
    sheet.sheet_view.rightToLeft = True


def style_header_row(sheet) -> None:
    """Apply simple header styling."""
    fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def auto_size_columns(sheet) -> None:
    """Best-effort column sizing for readable XLSX output."""
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def cell_value(value: Any) -> Any:
    """Normalize cell values while preserving numbers and booleans."""
    if value is None:
        return MISSING_FIELD_LABEL
    if isinstance(value, str) and not value.strip():
        return MISSING_FIELD_LABEL
    return value
