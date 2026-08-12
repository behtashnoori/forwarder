"""MT-1C.2 full-surface certification on disposable PostgreSQL 18.

The suite deliberately separates bootstrap, publisher, and request/reader
sessions.  It exercises product entry points; low-level ORM assertions are
included only as supporting evidence for composite and transaction fences.
"""

from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta, timezone
from hashlib import sha256
from io import BytesIO
import os
from pathlib import Path
from threading import Event

import pytest
from alembic import command
from alembic.config import Config
from openpyxl import load_workbook
from sqlalchemy import create_engine, delete, select, text, update
from sqlalchemy.orm import Session

from backend import create_app
from backend.census_context import ensure_census_context
from backend.extensions import db
from backend.models import (
    CaseDocumentFile,
    CaseDocumentRequirement,
    Customer,
    DocumentDefinition,
    ExpertConsoleLog,
    ExpertConsoleNotification,
    ExpertUser,
    ReferralAssignmentLog,
    ShipmentRequest,
)
from backend.operational_models import (
    CanonicalLocation,
    Milestone,
    OperationalAudit,
    OperationalOutbox,
    OperationalMembership,
    OperationalOrganization,
    OperationalShipment,
    OperationalWorkItem,
    RouteLeg,
    RoutePlan,
    Project,
    project_party_relationship,
)
from backend.ownership_census import (
    CensusDecisionInput,
    CensusPublication,
    internal_publisher_authority,
    publish_census,
    OwnershipActiveCensus,
    OwnershipDecision,
)
from backend.resource_identity import (
    ResourceIdentity,
    project_party_identity,
    scalar_identity,
)
from backend.services.auth_session_service import create_session_tokens
from backend.quarantine import QuarantinedResource, assert_instance_current
from backend.referral_engine import ReferralEngine
from backend.services import case_document_service, operational_service
from backend.services.document_storage_service import PrivateDocumentStorage


FP = sha256(b"mt1c2-full-surface-postgresql").hexdigest()
TOKEN = "mt1c2-full-surface-publisher-token"
UNSAFE = {
    "QUARANTINED": ("DETERMINISTIC", "QUARANTINED"),
    "CONFLICT": ("CONFLICT", "QUARANTINED"),
    "UNRESOLVED": ("UNRESOLVED", "QUARANTINED"),
    "INVALID_LINEAGE": ("INVALID_LINEAGE", "QUARANTINED"),
}


def _url() -> str:
    url = os.getenv("MT1C2_DISPOSABLE_DATABASE_URL", "")
    if not url:
        pytest.skip("explicit disposable MT-1C.2 PostgreSQL URL not provided")
    assert "127.0.0.1" in url or "localhost" in url
    assert "/forwarder_mt1c2_cert_" in url
    return url


def _decision(identity, classification="DETERMINISTIC", enforcement="CLEAR", root=None):
    return CensusDecisionInput(identity, classification, enforcement, FP, root)


def _publication(census_id, order, decisions, previous=None):
    counts = {}
    for item in decisions:
        counts[item.identity.resource_type] = (
            counts.get(item.identity.resource_type, 0) + 1
        )
    return CensusPublication(
        census_id,
        "mt1c2-pg-v1",
        order,
        previous,
        FP,
        "pytest-publisher",
        tuple(decisions),
        counts,
        {name: FP for name in counts},
    )


def _active_decisions(session):
    """Carry the active immutable census forward before adding a surface slice."""
    active = session.execute(
        select(OwnershipActiveCensus.census_id).execution_options(
            include_quarantined_for_certification=True
        )
    ).scalar_one()
    rows = session.execute(
        select(OwnershipDecision)
        .where(OwnershipDecision.census_id == active)
        .execution_options(include_quarantined_for_certification=True)
    ).scalars()
    return [
        _decision(
            ResourceIdentity.from_payload(row.resource_type, row.resource_key_payload),
            row.classification,
            row.enforcement_state,
            ResourceIdentity.from_payload(
                row.root_resource_type, row.root_resource_key_payload
            ),
        )
        for row in rows
    ]


@pytest.fixture(scope="module")
def certified_app():
    url = _url()
    os.environ["MT1D_CENSUS_PUBLISHER_TOKEN"] = TOKEN
    os.environ["MT1D_CENSUS_PUBLISHER_DATABASE_ROLES"] = "postgres"
    config = Config("backend/migrations/alembic.ini")
    config.set_main_option("script_location", "backend/migrations")
    config.set_main_option("sqlalchemy.url", url.replace("%", "%%"))
    engine = create_engine(url)
    with engine.begin() as connection:
        connection.execute(
            text(
                "CREATE TABLE IF NOT EXISTS alembic_version "
                "(version_num VARCHAR(255) NOT NULL PRIMARY KEY)"
            )
        )
    command.upgrade(config, "head")
    app = create_app(
        {
            "TESTING": True,
            "SQLALCHEMY_DATABASE_URI": url,
            "SECRET_KEY": "mt1c2-postgresql-only",
        },
        skip_startup=True,
    )
    try:
        with app.app_context():
            assert (
                db.session.execute(text("SHOW server_version"))
                .scalar_one()
                .startswith("18.")
            )
        yield app, engine
    finally:
        engine.dispose()


def test_01_real_detail_list_search_report_aggregate_export_pagination_and_tracking(
    certified_app,
):
    """Surfaces 1-8 and 15, including all fail-closed decision states."""
    app, engine = certified_app
    markers = ["CLEAR", *UNSAFE]
    with Session(engine) as bootstrap, bootstrap.begin():
        admin = ExpertUser(
            username="mt1c2-admin",
            password_hash="x",
            full_name="MT1C2 Admin",
            email="mt1c2-admin@example.test",
            role="admin",
            is_active=True,
        )
        bootstrap.add(admin)
        bootstrap.flush()
        organization = OperationalOrganization(name="MT1C2 certification tenant")
        bootstrap.add(organization)
        bootstrap.flush()
        bootstrap.add(OperationalMembership(
            organization_id=organization.id,
            user_id=admin.id,
            permissions=["operational_shipment.create_direct", "work_item.read", "work_item.manage"],
        ))
        rows = []
        for marker in markers:
            rows.append(
                ShipmentRequest(
                    ownership_scope="TENANT",
                    operational_organization_id=organization.id,
                    tracking_code=f"MT1C2-{marker}",
                    contact_phone=f"surface-{marker}",
                    customer_first_name=marker,
                    customer_last_name="Matrix",
                    shipping_type="domestic",
                    transport_method="road",
                    status="new",
                    status_request_status="new",
                    assigned_to=admin.id,
                )
            )
        bootstrap.add_all(rows)
        bootstrap.flush()
        admin_id = admin.id
        organization_id = organization.id
        ids = {marker: row.id for marker, row in zip(markers, rows)}

    decisions = [_decision(scalar_identity("ShipmentRequest", ids["CLEAR"]))]
    decisions.extend(
        _decision(
            scalar_identity("ShipmentRequest", ids[name]), classification, enforcement
        )
        for name, (classification, enforcement) in UNSAFE.items()
    )
    with Session(engine) as publisher:
        publish_census(
            publisher,
            _publication("mt1c2-surface-n1", 1, decisions),
            authority=internal_publisher_authority(TOKEN),
        )
    # A row created after the certified scope became active has no decision in
    # N and must fail closed for every reader pinned to N.
    with Session(engine) as bootstrap, bootstrap.begin():
        missing = ShipmentRequest(
            ownership_scope="TENANT",
            operational_organization_id=organization_id,
            tracking_code="MT1C2-MISSING_METADATA",
            contact_phone="surface-MISSING_METADATA",
            customer_first_name="MISSING_METADATA",
            customer_last_name="Matrix",
            shipping_type="domestic",
            transport_method="road",
            status="new",
            status_request_status="new",
            assigned_to=admin_id,
        )
        bootstrap.add(missing)
        bootstrap.flush()
        ids["MISSING_METADATA"] = missing.id

    with app.app_context():
        token = create_session_tokens(admin_id)["access_token"]
    headers = {"Authorization": f"Bearer {token}"}
    client = app.test_client()

    clear_detail = client.get(
        f"/api/admin/shipment-requests/{ids['CLEAR']}", headers=headers
    )
    assert clear_detail.status_code == 200
    assert clear_detail.get_json()["id"] == ids["CLEAR"]
    for marker in [*UNSAFE, "MISSING_METADATA"]:
        assert (
            client.get(
                f"/api/admin/shipment-requests/{ids[marker]}", headers=headers
            ).status_code
            == 404
        )

    listing = client.get(
        "/api/admin/shipment-requests?limit=2&offset=0", headers=headers
    )
    assert listing.status_code == 200
    payload = listing.get_json()
    serialized = str(payload)
    assert ids["CLEAR"] in [item["id"] for item in payload["requests"]]
    assert payload["pagination"]["total"] == 1
    assert not any(marker in serialized for marker in [*UNSAFE, "MISSING_METADATA"])

    search = client.get(
        "/api/expert/requests?search=surface-&per_page=2", headers=headers
    )
    assert search.status_code == 200
    search_payload = search.get_json()
    assert [row["id"] for row in search_payload["requests"]] == [ids["CLEAR"]]
    assert search_payload["pagination"]["total"] == 1

    dashboard = client.get("/api/admin/dashboard", headers=headers)
    assert dashboard.status_code == 200
    assert dashboard.get_json()["total_requests"] == 1
    report = client.get("/api/admin/reports/overview?period=weekly", headers=headers)
    assert report.status_code == 200
    report_text = report.get_data(as_text=True)
    assert "CLEAR" in report_text
    assert not any(marker in report_text for marker in [*UNSAFE, "MISSING_METADATA"])

    exported = client.get(
        "/api/admin/reports/export.xlsx?period=weekly", headers=headers
    )
    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.data), read_only=True, data_only=True)
    workbook_text = " ".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "CLEAR" in workbook_text
    assert not any(marker in workbook_text for marker in [*UNSAFE, "MISSING_METADATA"])

    assert client.get("/api/public/track/MT1C2-CLEAR").status_code == 200
    missing_body = client.get("/api/public/track/not-present").get_json()
    for marker in [*UNSAFE, "MISSING_METADATA"]:
        denied = client.get(f"/api/public/track/MT1C2-{marker}")
        assert denied.status_code == 404
        assert denied.get_json() == missing_body
    # Flask's scoped request session must not retain the shared publication
    # lock between independent certification scenarios.
    with app.app_context():
        db.session.remove()


def test_02_composite_select_update_delete_insert_and_pinned_n_plus_one(certified_app):
    """Surfaces 8 and 14: Core composite fencing and N/N+1 consistency."""
    app, engine = certified_app
    with Session(engine) as bootstrap, bootstrap.begin():
        admin = bootstrap.execute(
            select(ExpertUser).where(ExpertUser.username == "mt1c2-admin")
        ).scalar_one()
        organization = bootstrap.execute(
            select(OperationalOrganization).where(
                OperationalOrganization.name == "MT1C2 certification tenant"
            )
        ).scalar_one()
        from backend.models import Customer

        customer = Customer(
            company_name="MT1C2 composite", first_name="Composite", last_name="Customer",
            ownership_scope="TENANT", operational_organization_id=organization.id,
        )
        bootstrap.add(customer)
        bootstrap.flush()
        project = Project(
            organization_id=organization.id,
            primary_customer_id=customer.id,
            project_code="MT1C2-COMPOSITE",
            created_by_user_id=admin.id,
        )
        bootstrap.add(project)
        bootstrap.flush()
        bootstrap.execute(
            project_party_relationship.insert()
            .values(
                project_id=project.id,
                customer_id=customer.id,
                party_role="payer",
                source="clear",
            )
            .execution_options(include_quarantined_for_certification=True)
        )
        project_id, customer_id = project.id, customer.id

    project_identity = scalar_identity("Project", project_id)
    customer_identity = scalar_identity("Customer", customer_id)
    party_identity = project_party_identity(project_id, customer_id, "payer")
    with Session(engine) as census_reader:
        request_ids = (
            census_reader.execute(
                select(ShipmentRequest.id).execution_options(
                    include_quarantined_for_certification=True
                )
            )
            .scalars()
            .all()
        )
    request_decisions = [
        _decision(scalar_identity("ShipmentRequest", request_id))
        for request_id in request_ids
    ]
    with Session(engine) as publisher:
        publish_census(
            publisher,
            _publication(
                "mt1c2-composite-n2",
                2,
                [
                    *request_decisions,
                    _decision(project_identity),
                    _decision(customer_identity),
                    _decision(party_identity, root=project_identity),
                ],
                previous="mt1c2-surface-n1",
            ),
            authority=internal_publisher_authority(TOKEN),
        )

    pinned = Event()
    release = Event()
    published = Event()

    def reader():
        with Session(engine) as session, session.begin():
            context = ensure_census_context(session)
            assert session.execute(select(project_party_relationship)).all()
            pinned.set()
            assert release.wait(10)
            assert ensure_census_context(session) == context
            assert session.execute(select(project_party_relationship)).all()

    def publisher():
        assert pinned.wait(10)
        with Session(engine) as session:
            publish_census(
                session,
                _publication(
                    "mt1c2-composite-n3",
                    3,
                    [
                        *request_decisions,
                        _decision(project_identity, "CONFLICT", "QUARANTINED"),
                        _decision(customer_identity),
                        _decision(party_identity, root=project_identity),
                    ],
                    previous="mt1c2-composite-n2",
                ),
                authority=internal_publisher_authority(TOKEN),
            )
        published.set()

    with ThreadPoolExecutor(max_workers=2) as pool:
        read_future = pool.submit(reader)
        publish_future = pool.submit(publisher)
        assert pinned.wait(10)
        assert not published.wait(0.5)
        release.set()
        read_future.result(timeout=15)
        publish_future.result(timeout=15)

    with Session(engine) as session, session.begin():
        assert session.execute(select(project_party_relationship)).all() == []
        assert (
            session.execute(
                update(project_party_relationship).values(source="bypass")
            ).rowcount
            == 0
        )
        assert session.execute(delete(project_party_relationship)).rowcount == 0
        with pytest.raises(Exception):
            session.execute(
                project_party_relationship.insert().values(
                    project_id=project_id,
                    customer_id=customer_id,
                    party_role="notify_party",
                    source="unsafe-insert",
                )
            )


def test_03_real_customer_and_joined_project_selectors_six_state(certified_app):
    """Surface 4: actual autocomplete routes, including joined-parent roots."""
    app, engine = certified_app
    markers = ["CLEAR", *UNSAFE]
    with Session(engine) as bootstrap, bootstrap.begin():
        admin = bootstrap.execute(
            select(ExpertUser).where(ExpertUser.username == "mt1c2-admin")
        ).scalar_one()
        organization = bootstrap.execute(
            select(OperationalOrganization).where(
                OperationalOrganization.name == "MT1C2 certification tenant"
            )
        ).scalar_one()
        customers = []
        project_ids = {}
        for marker in markers:
            customer = Customer(
                ownership_scope="TENANT",
                operational_organization_id=organization.id,
                company_name=f"MT1C2-SELECTOR-{marker}",
                first_name=marker,
                last_name="Selector",
                status="active",
            )
            bootstrap.add(customer)
            bootstrap.flush()
            project_id = bootstrap.execute(
                Project.__table__.insert()
                .values(
                    organization_id=organization.id,
                    primary_customer_id=customer.id,
                    project_code=f"MT1C2-PROJECT-{marker}",
                    created_by_user_id=admin.id,
                )
                .returning(Project.id)
                .execution_options(include_quarantined_for_certification=True)
            ).scalar_one()
            customers.append(customer)
            project_ids[marker] = project_id
        admin_id = admin.id
        organization_id = organization.id
        customer_ids = {m: row.id for m, row in zip(markers, customers)}

    with Session(engine) as reader:
        decisions = _active_decisions(reader)
    for marker in markers:
        classification, enforcement = UNSAFE.get(marker, ("DETERMINISTIC", "CLEAR"))
        customer_identity = scalar_identity("Customer", customer_ids[marker])
        decisions.extend(
            [
                _decision(customer_identity, classification, enforcement),
                _decision(
                    scalar_identity("Project", project_ids[marker]),
                    root=customer_identity,
                ),
            ]
        )
    with Session(engine) as publisher:
        publish_census(
            publisher,
            _publication("mt1c2-selector-n4", 4, decisions, "mt1c2-composite-n3"),
            authority=internal_publisher_authority(TOKEN),
        )

    # Post-publication rows intentionally lack certification metadata.
    with Session(engine) as bootstrap, bootstrap.begin():
        missing_customer = Customer(
            ownership_scope="TENANT",
            operational_organization_id=organization_id,
            company_name="MT1C2-SELECTOR-MISSING_METADATA",
            first_name="Missing",
            last_name="Selector",
            status="active",
        )
        bootstrap.add(missing_customer)
        bootstrap.flush()
        bootstrap.execute(
            Project.__table__.insert()
            .values(
                organization_id=organization_id,
                primary_customer_id=missing_customer.id,
                project_code="MT1C2-PROJECT-MISSING_METADATA",
                created_by_user_id=admin_id,
            )
            .execution_options(include_quarantined_for_certification=True)
        )

    with app.app_context():
        token = create_session_tokens(admin_id)["access_token"]
        db.session.remove()
    headers = {"Authorization": f"Bearer {token}"}
    client = app.test_client()
    customers_response = client.get(
        "/api/operations/selectors/customers?q=MT1C2-SELECTOR-&limit=100",
        headers=headers,
    )
    assert customers_response.status_code == 200
    assert customers_response.get_json()["items"] == [
        {"id": customer_ids["CLEAR"], "label": "MT1C2-SELECTOR-CLEAR"}
    ]
    projects_response = client.get(
        "/api/operations/selectors/projects?q=MT1C2-PROJECT-&limit=100",
        headers=headers,
    )
    assert projects_response.status_code == 200
    assert [item["label"] for item in projects_response.get_json()["items"]] == [
        "MT1C2-PROJECT-CLEAR"
    ]


def test_04_real_reconciliation_job_outbox_and_publisher_fence(
    certified_app, monkeypatch
):
    """Surfaces 9 and 12: real overdue job, atomic audit/outbox, and N/N+1 fence."""
    app, engine = certified_app
    now = datetime.now(timezone.utc)
    with Session(engine) as bootstrap, bootstrap.begin():
        admin = bootstrap.scalar(
            select(ExpertUser).where(ExpertUser.username == "mt1c2-admin")
        )
        org = bootstrap.execute(
            select(OperationalOrganization).where(
                OperationalOrganization.name == "MT1C2 certification tenant"
            )
        ).scalar_one()
        customer = Customer(
            ownership_scope="TENANT",
            operational_organization_id=org.id,
            company_name="MT1C2 job customer",
            first_name="Job",
            last_name="Customer",
            status="active",
        )
        bootstrap.add(customer)
        bootstrap.flush()
        locations = [
            CanonicalLocation(
                source_type="province",
                source_id=91001 + i,
                location_type="province",
                display_name=f"Job {i}",
            )
            for i in range(2)
        ]
        bootstrap.add_all(locations)
        bootstrap.flush()
        shipments = []
        for marker in ("CLEAR", "QUARANTINED"):
            shipment_id = bootstrap.execute(
                OperationalShipment.__table__.insert()
                .values(
                    organization_id=org.id,
                    source_type="direct",
                    customer_id=customer.id,
                    lifecycle_status="planned",
                    created_by_user_id=admin.id,
                )
                .returning(OperationalShipment.id)
                .execution_options(include_quarantined_for_certification=True)
            ).scalar_one()
            plan = RoutePlan(
                operational_shipment_id=shipment_id,
                revision=1,
                status="active",
                is_active=True,
                created_by_user_id=admin.id,
            )
            bootstrap.add(plan)
            bootstrap.flush()
            leg = RouteLeg(
                route_plan_id=plan.id,
                sequence_number=1,
                origin_location_id=locations[0].id,
                destination_location_id=locations[1].id,
                origin_snapshot={"name": "A"},
                destination_snapshot={"name": "B"},
                transport_mode="road",
                planned_departure=now - timedelta(hours=3),
                planned_arrival=now - timedelta(hours=2),
            )
            bootstrap.add(leg)
            bootstrap.flush()
            bootstrap.add(
                Milestone(
                    organization_id=org.id,
                    operational_shipment_id=shipment_id,
                    route_plan_id=plan.id,
                    route_leg_id=leg.id,
                    milestone_type=f"arrival-{marker.lower()}",
                    planned_at=now - timedelta(hours=1),
                )
            )
            shipments.append(shipment_id)
        org_id, user_id, _customer_id = org.id, admin.id, customer.id
        shipment_ids = dict(zip(("CLEAR", "QUARANTINED"), shipments))

    with Session(engine) as reader:
        decisions = _active_decisions(reader)
        known = {item.identity for item in decisions}
        for row_id in reader.scalars(
            select(Customer.id).execution_options(
                include_quarantined_for_certification=True
            )
        ):
            identity = scalar_identity("Customer", row_id)
            if identity not in known:
                decisions.append(_decision(identity))
                known.add(identity)
        for project_row in reader.execute(
            select(Project.id, Project.primary_customer_id).execution_options(
                include_quarantined_for_certification=True
            )
        ):
            identity = scalar_identity("Project", project_row.id)
            if identity not in known:
                decisions.append(
                    _decision(
                        identity,
                        root=scalar_identity(
                            "Customer", project_row.primary_customer_id
                        ),
                    )
                )
                known.add(identity)
    decisions.extend(
        [
            _decision(scalar_identity("OperationalShipment", shipment_ids["CLEAR"])),
            _decision(
                scalar_identity("OperationalShipment", shipment_ids["QUARANTINED"]),
                "DETERMINISTIC",
                "QUARANTINED",
            ),
        ]
    )
    with Session(engine) as publisher:
        publish_census(
            publisher,
            _publication("mt1c2-job-n5", 5, decisions, "mt1c2-selector-n4"),
            authority=internal_publisher_authority(TOKEN),
        )

    pinned, release, published = Event(), Event(), Event()
    real_mutations = operational_service._reconcile_overdue_mutations

    def paused_mutations(*args, **kwargs):
        pinned.set()
        assert release.wait(10)
        return real_mutations(*args, **kwargs)

    monkeypatch.setattr(
        operational_service, "_reconcile_overdue_mutations", paused_mutations
    )

    def run_job():
        with app.app_context():
            return operational_service.reconcile_overdue(
                user_id=user_id, organization_id=org_id, now=now
            )

    next_decisions = [
        d
        for d in decisions
        if d.identity != scalar_identity("OperationalShipment", shipment_ids["CLEAR"])
    ]
    next_decisions.append(
        _decision(
            scalar_identity("OperationalShipment", shipment_ids["CLEAR"]),
            "DETERMINISTIC",
            "QUARANTINED",
        )
    )

    def publish_next():
        with Session(engine) as publisher:
            publish_census(
                publisher,
                _publication("mt1c2-job-n6", 6, next_decisions, "mt1c2-job-n5"),
                authority=internal_publisher_authority(TOKEN),
            )
        published.set()

    with ThreadPoolExecutor(max_workers=2) as pool:
        job = pool.submit(run_job)
        assert pinned.wait(10)
        publication = pool.submit(publish_next)
        assert not published.wait(0.5)
        release.set()
        assert job.result(timeout=15) == 1
        publication.result(timeout=15)

    monkeypatch.setattr(
        operational_service, "_reconcile_overdue_mutations", real_mutations
    )
    with app.app_context():
        assert (
            operational_service.reconcile_overdue(
                user_id=user_id, organization_id=org_id, now=now
            )
            == 0
        )
        assert OperationalWorkItem.query.count() == 1
        assert OperationalAudit.query.filter_by(action="work_item.opened").count() == 1
        assert (
            OperationalOutbox.query.filter_by(event_type="work_item.opened").count()
            == 1
        )
        db.session.remove()


def test_05_real_assignment_and_notification_six_state_atomicity(certified_app):
    """Surfaces 10 and 11: referral assignment and its notification share caller commit."""
    app, engine = certified_app
    markers = ["CLEAR", *UNSAFE]
    with Session(engine) as bootstrap, bootstrap.begin():
        expert = ExpertUser(
            username="mt1c2-assignee",
            password_hash="x",
            full_name="Assignee",
            email="mt1c2-assignee@example.test",
            role="expert",
            is_active=True,
        )
        bootstrap.add(expert)
        bootstrap.flush()
        organization = bootstrap.execute(
            select(OperationalOrganization).where(
                OperationalOrganization.name == "MT1C2 certification tenant"
            )
        ).scalar_one()
        bootstrap.add(OperationalMembership(
            organization_id=organization.id, user_id=expert.id, permissions=[]
        ))
        requests = []
        for marker in markers:
            row = ShipmentRequest(
                ownership_scope="TENANT",
                operational_organization_id=organization.id,
                tracking_code=f"MT1C2-ASSIGN-{marker}",
                contact_phone=f"assign-{marker}",
                shipping_type="domestic",
                transport_method="road",
                status="new",
                status_request_status="new",
            )
            bootstrap.add(row)
            bootstrap.flush()
            requests.append(row)
        request_ids = {m: r.id for m, r in zip(markers, requests)}
        expert_id = expert.id
        organization_id = organization.id
    with Session(engine) as reader:
        decisions = _active_decisions(reader)
    for marker in ["CLEAR", *UNSAFE]:
        classification, enforcement = UNSAFE.get(marker, ("DETERMINISTIC", "CLEAR"))
        decisions.append(
            _decision(
                scalar_identity("ShipmentRequest", request_ids[marker]),
                classification,
                enforcement,
            )
        )
    with Session(engine) as publisher:
        publish_census(
            publisher,
            _publication("mt1c2-assign-n7", 7, decisions, "mt1c2-job-n6"),
            authority=internal_publisher_authority(TOKEN),
        )
    with Session(engine) as bootstrap, bootstrap.begin():
        missing = ShipmentRequest(
            ownership_scope="TENANT",
            operational_organization_id=organization_id,
            tracking_code="MT1C2-ASSIGN-MISSING_METADATA",
            contact_phone="assign-MISSING_METADATA",
            shipping_type="domestic",
            transport_method="road",
            status="new",
            status_request_status="new",
        )
        bootstrap.add(missing)
        bootstrap.flush()
        request_ids["MISSING_METADATA"] = missing.id

    with app.app_context():
        with db.session.begin():
            assert (
                ReferralEngine(db.session).auto_assign_request(request_ids["CLEAR"])
                == expert_id
            )
        db.session.remove()
    with Session(engine) as check:
        assert (
            check.execute(
                select(ShipmentRequest.assigned_to)
                .where(ShipmentRequest.id == request_ids["CLEAR"])
                .execution_options(include_quarantined_for_certification=True)
            ).scalar_one()
            == expert_id
        )
        assert (
            check.scalar(
                select(ReferralAssignmentLog.id)
                .where(ReferralAssignmentLog.request_id == request_ids["CLEAR"])
                .execution_options(include_quarantined_for_certification=True)
            )
            is not None
        )
        assert (
            check.scalar(
                select(ExpertConsoleNotification.id)
                .where(
                    ExpertConsoleNotification.shipment_request_id
                    == request_ids["CLEAR"]
                )
                .execution_options(include_quarantined_for_certification=True)
            )
            is not None
        )
    for marker in [*UNSAFE, "MISSING_METADATA"]:
        with app.app_context():
            with pytest.raises(QuarantinedResource):
                with db.session.begin():
                    ReferralEngine(db.session).auto_assign_request(request_ids[marker])
            db.session.rollback()
            db.session.remove()
        with Session(engine) as check:
            opts = {"include_quarantined_for_certification": True}
            assert (
                check.execute(
                    select(ShipmentRequest.assigned_to)
                    .where(ShipmentRequest.id == request_ids[marker])
                    .execution_options(**opts)
                ).scalar_one()
                is None
            )
            assert (
                check.execute(
                    select(ReferralAssignmentLog.id)
                    .where(ReferralAssignmentLog.request_id == request_ids[marker])
                    .execution_options(**opts)
                ).all()
                == []
            )
            assert (
                check.execute(
                    select(ExpertConsoleNotification.id)
                    .where(
                        ExpertConsoleNotification.shipment_request_id
                        == request_ids[marker]
                    )
                    .execution_options(**opts)
                ).all()
                == []
            )
            assert (
                check.execute(
                    select(ExpertConsoleLog.id)
                    .where(ExpertConsoleLog.shipment_request_id == request_ids[marker])
                    .execution_options(**opts)
                ).all()
                == []
            )


def test_06_real_document_metadata_download_and_descendant_transition(certified_app):
    """Surfaces 13-14: route metadata/download plus held root-child-descendant revalidation."""
    app, engine = certified_app
    markers = ["CLEAR", *UNSAFE]
    storage_root = Path(app.config["DOCUMENT_STORAGE_ROOT"])
    with Session(engine) as bootstrap, bootstrap.begin():
        admin = bootstrap.scalar(
            select(ExpertUser).where(ExpertUser.username == "mt1c2-admin")
        )
        definition = DocumentDefinition(
            code="MT1C2-DOC",
            title="MT1C2 document",
            is_required=True,
            allowed_formats='["pdf"]',
            max_file_size_bytes=1024,
            max_active_file_count=1,
            sort_order=1,
        )
        bootstrap.add(definition)
        bootstrap.flush()
        organization = bootstrap.execute(
            select(OperationalOrganization).where(
                OperationalOrganization.name == "MT1C2 certification tenant"
            )
        ).scalar_one()
        records = {}
        for marker in markers:
            case = ShipmentRequest(
                ownership_scope="TENANT",
                operational_organization_id=organization.id,
                tracking_code=f"MT1C2-DOC-{marker}",
                contact_phone=f"doc-{marker}",
                shipping_type="domestic",
                transport_method="road",
                status="new",
                status_request_status="new",
                assigned_to=admin.id,
            )
            bootstrap.add(case)
            bootstrap.flush()
            requirement_id = bootstrap.execute(
                CaseDocumentRequirement.__table__.insert()
                .values(
                    shipment_request_id=case.id,
                    source_definition_id=definition.id,
                    source_definition_code=definition.code,
                    source_definition_revision=1,
                    title=definition.title,
                    is_required=True,
                    allowed_formats='["pdf"]',
                    max_file_size_bytes=1024,
                    max_active_file_count=1,
                    sort_order=1,
                )
                .returning(CaseDocumentRequirement.id)
                .execution_options(include_quarantined_for_certification=True)
            ).scalar_one()
            key = f"{case.id}/aa/{marker.lower()}.pdf"
            file_id = bootstrap.execute(
                CaseDocumentFile.__table__.insert()
                .values(
                    shipment_request_id=case.id,
                    case_requirement_id=requirement_id,
                    is_miscellaneous=False,
                    original_filename=f"{marker}.pdf",
                    safe_download_filename=f"{marker}.pdf",
                    storage_key=key,
                    canonical_extension="pdf",
                    detected_mime_type="application/pdf",
                    file_size_bytes=9,
                    sha256_hash="a" * 64,
                    version_number=1,
                    status="active",
                    uploaded_by=admin.id,
                )
                .returning(CaseDocumentFile.id)
                .execution_options(include_quarantined_for_certification=True)
            ).scalar_one()
            records[marker] = (case.id, requirement_id, file_id, key)
        admin_id, definition_id = admin.id, definition.id
        organization_id = organization.id
    for marker, (_, _, _, key) in records.items():
        path = storage_root / key
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(b"pdf-bytes")

    with Session(engine) as reader:
        decisions = _active_decisions(reader)
        known = {item.identity for item in decisions}
        document_case_ids = {value[0] for value in records.values()}
        for request_id in reader.scalars(
            select(ShipmentRequest.id).execution_options(
                include_quarantined_for_certification=True
            )
        ):
            identity = scalar_identity("ShipmentRequest", request_id)
            if request_id not in document_case_ids and identity not in known:
                decisions.append(_decision(identity))
                known.add(identity)
    for marker in ["CLEAR", *UNSAFE]:
        case_id, requirement_id, file_id, _ = records[marker]
        classification, enforcement = UNSAFE.get(marker, ("DETERMINISTIC", "CLEAR"))
        root = scalar_identity("ShipmentRequest", case_id)
        decisions.extend(
            [
                _decision(root, classification, enforcement),
                _decision(
                    scalar_identity("CaseDocumentRequirement", requirement_id),
                    classification,
                    enforcement,
                    root=root,
                ),
                _decision(
                    scalar_identity("CaseDocumentFile", file_id),
                    classification,
                    enforcement,
                    root=root,
                ),
            ]
        )
    with Session(engine) as publisher:
        publish_census(
            publisher,
            _publication("mt1c2-doc-n8", 8, decisions, "mt1c2-assign-n7"),
            authority=internal_publisher_authority(TOKEN),
        )
    with Session(engine) as bootstrap, bootstrap.begin():
        case = ShipmentRequest(
            ownership_scope="TENANT",
            operational_organization_id=organization_id,
            tracking_code="MT1C2-DOC-MISSING_METADATA",
            contact_phone="doc-MISSING_METADATA",
            shipping_type="domestic",
            transport_method="road",
            status="new",
            status_request_status="new",
            assigned_to=admin_id,
        )
        bootstrap.add(case)
        bootstrap.flush()
        requirement_id = bootstrap.execute(
            CaseDocumentRequirement.__table__.insert()
            .values(
                shipment_request_id=case.id,
                source_definition_id=definition_id,
                source_definition_code="MT1C2-DOC",
                source_definition_revision=1,
                title="MT1C2 document",
                is_required=True,
                allowed_formats='["pdf"]',
                max_file_size_bytes=1024,
                max_active_file_count=1,
                sort_order=1,
            )
            .returning(CaseDocumentRequirement.id)
            .execution_options(include_quarantined_for_certification=True)
        ).scalar_one()
        key = f"{case.id}/aa/missing_metadata.pdf"
        file_id = bootstrap.execute(
            CaseDocumentFile.__table__.insert()
            .values(
                shipment_request_id=case.id,
                case_requirement_id=requirement_id,
                is_miscellaneous=False,
                original_filename="MISSING_METADATA.pdf",
                safe_download_filename="MISSING_METADATA.pdf",
                storage_key=key,
                canonical_extension="pdf",
                detected_mime_type="application/pdf",
                file_size_bytes=9,
                sha256_hash="a" * 64,
                version_number=1,
                status="active",
                uploaded_by=admin_id,
            )
            .returning(CaseDocumentFile.id)
            .execution_options(include_quarantined_for_certification=True)
        ).scalar_one()
        records["MISSING_METADATA"] = (case.id, requirement_id, file_id, key)
    missing_path = storage_root / records["MISSING_METADATA"][3]
    missing_path.parent.mkdir(parents=True, exist_ok=True)
    missing_path.write_bytes(b"pdf-bytes")
    with app.app_context():
        token = create_session_tokens(admin_id)["access_token"]
        db.session.remove()
    headers = {"Authorization": f"Bearer {token}"}
    client = app.test_client()
    clear_case, clear_requirement, clear_file, clear_key = records["CLEAR"]
    metadata = client.get(
        f"/api/expert/requests/{clear_case}/documents", headers=headers
    )
    assert metadata.status_code == 200
    assert metadata.get_json()["requirements"][0]["active_files"][0]["id"] == clear_file
    download = client.get(
        f"/api/expert/requests/{clear_case}/documents/{clear_file}/download",
        headers=headers,
    )
    assert download.status_code == 200 and download.data == b"pdf-bytes"
    for marker in [*UNSAFE, "MISSING_METADATA"]:
        case_id, _, file_id, _ = records[marker]
        assert (
            client.get(
                f"/api/expert/requests/{case_id}/documents", headers=headers
            ).status_code
            == 404
        )
        assert (
            client.get(
                f"/api/expert/requests/{case_id}/documents/{file_id}/download",
                headers=headers,
            ).status_code
            == 404
        )
    other_case, _, other_file, _ = records["QUARANTINED"]
    assert (
        client.get(
            f"/api/expert/requests/{clear_case}/documents/{other_file}/download",
            headers=headers,
        ).status_code
        == 404
    )
    with app.app_context():
        assert not hasattr(PrivateDocumentStorage(), "resolve")
        held_case = db.session.get(ShipmentRequest, clear_case)
        held_requirement = db.session.get(CaseDocumentRequirement, clear_requirement)
        held_file = db.session.get(CaseDocumentFile, clear_file)
        assert (
            PrivateDocumentStorage()
            .resolve_for_download(held_file, case=held_case)
            .is_file()
        )
        db.session.expunge_all()
        db.session.remove()
    next_decisions = []
    clear_lineage = {
        scalar_identity("ShipmentRequest", clear_case),
        scalar_identity("CaseDocumentRequirement", clear_requirement),
        scalar_identity("CaseDocumentFile", clear_file),
    }
    for decision in decisions:
        if decision.identity in clear_lineage:
            next_decisions.append(
                _decision(
                    decision.identity,
                    "DETERMINISTIC",
                    "QUARANTINED",
                    root=scalar_identity("ShipmentRequest", clear_case)
                    if decision.identity.resource_type != "ShipmentRequest"
                    else None,
                )
            )
        else:
            next_decisions.append(decision)
    missing_case, missing_requirement, missing_file, _ = records["MISSING_METADATA"]
    missing_root = scalar_identity("ShipmentRequest", missing_case)
    next_decisions.extend(
        [
            _decision(missing_root),
            _decision(
                scalar_identity("CaseDocumentRequirement", missing_requirement),
                root=missing_root,
            ),
            _decision(
                scalar_identity("CaseDocumentFile", missing_file), root=missing_root
            ),
        ]
    )
    with Session(engine) as publisher:
        publish_census(
            publisher,
            _publication("mt1c2-doc-n9", 9, next_decisions, "mt1c2-doc-n8"),
            authority=internal_publisher_authority(TOKEN),
        )
    with app.app_context():
        with pytest.raises(QuarantinedResource):
            case_document_service.case_payload(held_case)
        with pytest.raises(QuarantinedResource):
            case_document_service.serialize_file(held_file)
        with pytest.raises(QuarantinedResource):
            assert_instance_current(held_requirement, purpose="held-descendant")
        assert (
            client.get(
                f"/api/expert/requests/{clear_case}/documents", headers=headers
            ).status_code
            == 404
        )
        assert (
            client.get(
                f"/api/expert/requests/{clear_case}/documents/{clear_file}/download",
                headers=headers,
            ).status_code
            == 404
        )
    with app.app_context():
        db.session.remove()
