"""Characterization tests for admin panel read/report route contracts."""
from __future__ import annotations

from io import BytesIO
from datetime import datetime, timedelta

import bcrypt
import pytest
from openpyxl import load_workbook

from backend import create_app
from backend.extensions import db
from backend.models import (
    AssignmentLog,
    City,
    County,
    Customer,
    ExpertConsoleLog,
    ExpertUser,
    Province,
    ShipmentRequest,
)
from backend.security import security
from backend.services.auth_session_service import create_session_tokens


@pytest.fixture
def admin_panel_app():
    """App with isolated DB and admin panel report seed data."""
    app = create_app(
        {
            "TESTING": True,
            "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
            "SECRET_KEY": "test-secret",
        },
        skip_startup=True,
    )
    with app.app_context():
        db.create_all()
        password_hash = bcrypt.hashpw(b"test123", bcrypt.gensalt()).decode("utf-8")
        admin = ExpertUser(
            username="phase5m_admin",
            password_hash=password_hash,
            full_name="Phase 5M Admin",
            email="phase5m-admin@example.test",
            role="admin",
            authority="PLATFORM_ADMIN",
            is_active=True,
        )
        expert = ExpertUser(
            username="phase5m_expert",
            password_hash=password_hash,
            full_name="Phase 5M Expert",
            email="phase5m-expert@example.test",
            role="expert",
            is_active=True,
        )
        province = Province(code="THR", name_fa="Tehran")
        db.session.add_all([admin, expert, province])
        db.session.flush()
        county = County(name_fa="Tehran County", province_id=province.id)
        db.session.add(county)
        db.session.flush()
        city = City(name_fa="Tehran City", county_id=county.id, province_id=province.id)
        db.session.add(city)
        db.session.flush()

        now = datetime.utcnow()
        assigned_request = ShipmentRequest(
            tracking_code="AP-P5M-001",
            shipping_type="domestic",
            contact_phone="09123456789",
            customer_first_name="Assigned",
            customer_last_name="Customer",
            transport_method="legacy-road",
            domestic_transport_method="road",
            status_request_status="new",
            status="assigned",
            assigned_to=expert.id,
            origin_province_id=province.id,
            origin_county_id=county.id,
            origin_city_id=city.id,
            dest_province_id=province.id,
            dest_county_id=county.id,
            dest_city_id=city.id,
            priority="high",
            created_at=now - timedelta(hours=2),
            sla_due_at=now - timedelta(minutes=30),
        )
        won_request = ShipmentRequest(
            tracking_code="AP-P5M-002",
            shipping_type="domestic",
            contact_phone="09123456780",
            customer_first_name="Won",
            customer_last_name="Customer",
            transport_method="air",
            status_request_status="new",
            status="won",
            assigned_to=expert.id,
            origin_province_id=province.id,
            created_at=now - timedelta(days=2),
        )
        new_request = ShipmentRequest(
            tracking_code="AP-P5M-003",
            shipping_type="domestic",
            contact_phone="09123456781",
            customer_first_name="New",
            customer_last_name="Customer",
            transport_method="sea",
            status_request_status="new",
            status="new",
            assigned_to=None,
            created_at=now - timedelta(days=10),
        )
        db.session.add_all([assigned_request, won_request, new_request])
        db.session.flush()

        assignment_log = AssignmentLog(
            shipment_request_id=assigned_request.id,
            assigned_expert_id=expert.id,
            assignment_method="automatic",
            assignment_reason="Seed assignment",
            created_at=now - timedelta(hours=2),
        )
        response_log = ExpertConsoleLog(
            shipment_request_id=assigned_request.id,
            expert_user_id=expert.id,
            action="message_added",
            created_at=now - timedelta(hours=1),
        )
        db.session.add_all([assignment_log, response_log])
        db.session.commit()

        return {
            "app": app,
            "admin_token": create_session_tokens(admin.id)["access_token"],
            "expert_token": create_session_tokens(expert.id)["access_token"],
            "expert_id": expert.id,
            "assigned_request_id": assigned_request.id,
            "new_request_id": new_request.id,
            "province_id": province.id,
        }


def _auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
def admin_empty_report_app():
    """App with an admin user and no shipment requests for empty report coverage."""
    app = create_app(
        {
            "TESTING": True,
            "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
            "SECRET_KEY": "test-secret",
        },
        skip_startup=True,
    )
    with app.app_context():
        db.create_all()
        password_hash = bcrypt.hashpw(b"test123", bcrypt.gensalt()).decode("utf-8")
        admin = ExpertUser(
            username="phase12b_empty_admin",
            password_hash=password_hash,
            full_name="Phase 12B Empty Admin",
            email="phase12b-empty-admin@example.test",
            role="admin",
            authority="PLATFORM_ADMIN",
            is_active=True,
        )
        db.session.add(admin)
        db.session.commit()

        return {
            "app": app,
            "admin_token": create_session_tokens(admin.id)["access_token"],
        }


def _assert_report_exclusions(payload):
    serialized = str(payload)
    assert "ارجاع شده" not in serialized
    assert "ارجاع‌شده" not in serialized

    def walk(value):
        if isinstance(value, dict):
            for key, item in value.items():
                assert "sla" not in key.lower()
                assert "priority" not in key.lower()
                assert "deadline" not in key.lower()
                assert "overdue" not in key.lower()
                assert "due_soon" not in key.lower()
                walk(item)
        elif isinstance(value, list):
            for item in value:
                walk(item)

    walk(payload)


def _assert_workbook_exclusions(workbook):
    visible_values = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            visible_values.extend(str(value) for value in row if value is not None)
    serialized = " ".join(visible_values)
    assert "ارجاع شده" not in serialized
    assert "ارجاع‌شده" not in serialized
    assert "SLA" not in serialized
    assert "sla" not in serialized
    assert "اولویت" not in serialized
    assert "priority" not in serialized


def test_admin_shipment_request_detail_and_list_contract(admin_panel_app):
    """Admin request detail/list keep auth, filters, date errors, pagination, and location shape."""
    client = admin_panel_app["app"].test_client()
    admin_headers = _auth_headers(admin_panel_app["admin_token"])
    expert_headers = _auth_headers(admin_panel_app["expert_token"])
    expected_forbidden_payload = {
        "error": "\u062f\u0633\u062a\u0631\u0633\u06cc \u063a\u06cc\u0631\u0645\u062c\u0627\u0632",
        "required_roles": ["admin"],
        "user_role": "expert",
    }

    unauthenticated_detail = client.get(
        f"/api/admin/shipment-requests/{admin_panel_app['assigned_request_id']}"
    )
    assert unauthenticated_detail.status_code == 401
    assert unauthenticated_detail.get_json() == {"error": "Token is missing"}

    forbidden_detail = client.get(
        f"/api/admin/shipment-requests/{admin_panel_app['assigned_request_id']}",
        headers=expert_headers,
    )
    assert forbidden_detail.status_code == 403
    assert forbidden_detail.get_json() == expected_forbidden_payload

    unauthenticated_list = client.get("/api/admin/shipment-requests")
    assert unauthenticated_list.status_code == 401
    assert unauthenticated_list.get_json() == {"error": "Token is missing"}

    forbidden_list = client.get("/api/admin/shipment-requests", headers=expert_headers)
    assert forbidden_list.status_code == 403
    assert forbidden_list.get_json() == expected_forbidden_payload

    missing_detail = client.get("/api/admin/shipment-requests/999999", headers=admin_headers)
    assert missing_detail.status_code == 404
    assert missing_detail.get_json() == {"error": "\u062f\u0631\u062e\u0648\u0627\u0633\u062a \u0645\u0648\u0631\u062f\u0646\u0638\u0631 \u06cc\u0627\u0641\u062a \u0646\u0634\u062f"}

    detail = client.get(f"/api/admin/shipment-requests/{admin_panel_app['assigned_request_id']}", headers=admin_headers)
    assert detail.status_code == 200
    detail_payload = detail.get_json()
    assert set(detail_payload.keys()) == {
        "id",
        "contact_phone",
        "customer_first_name",
        "customer_last_name",
        "transport_method",
        "status",
        "priority",
        "assigned_to",
        "shipping_type",
        "origin",
        "destination",
        "iran_destination",
        "created_at",
        "sla_due_at",
    }
    assert detail_payload["assigned_to"] == {
        "id": admin_panel_app["expert_id"],
        "full_name": "Phase 5M Expert",
        "username": "phase5m_expert",
    }
    assert detail_payload["origin"] == {
        "province": "Tehran", "county": "Tehran County", "city": "Tehran City",
        "country": None, "international_city": None, "address": None,
    }
    assert detail_payload["iran_destination"] is None

    invalid_date = client.get("/api/admin/shipment-requests?date_from=not-a-date", headers=admin_headers)
    assert invalid_date.status_code == 400
    assert invalid_date.get_json() == {"error": "\u0641\u0631\u0645\u062a \u062a\u0627\u0631\u06cc\u062e \u0646\u0627\u0645\u0639\u062a\u0628\u0631 \u0627\u0633\u062a"}

    list_response = client.get(
        f"/api/admin/shipment-requests?status=assigned&province_id={admin_panel_app['province_id']}&limit=1&offset=0",
        headers=admin_headers,
    )
    assert list_response.status_code == 200
    list_payload = list_response.get_json()
    assert set(list_payload.keys()) == {"requests", "pagination"}
    assert list_payload["pagination"] == {"total": 1, "limit": 1, "offset": 0, "has_more": False}
    assert len(list_payload["requests"]) == 1
    assert list_payload["requests"][0]["id"] == admin_panel_app["assigned_request_id"]
    assert list_payload["requests"][0]["assigned_to"] == admin_panel_app["expert_id"]
    assert list_payload["requests"][0]["origin"] == {
        "province": "Tehran", "county": "Tehran County", "city": "Tehran City",
        "country": None, "international_city": None, "address": None,
    }


@pytest.mark.skip(reason="ADR-043: reporting/export remains fail-closed pending a companion decision")
def test_admin_dashboard_and_assignment_summary_contract(admin_panel_app):
    """Dashboard and assignment summary keep aggregate keys, counts, and report shape."""
    client = admin_panel_app["app"].test_client()
    admin_headers = _auth_headers(admin_panel_app["admin_token"])
    expert_headers = _auth_headers(admin_panel_app["expert_token"])

    unauthenticated_dashboard = client.get("/api/admin/dashboard")
    assert unauthenticated_dashboard.status_code == 401
    assert unauthenticated_dashboard.get_json() == {"error": "Token is missing"}

    forbidden_dashboard = client.get("/api/admin/dashboard", headers=expert_headers)
    assert forbidden_dashboard.status_code == 403
    assert forbidden_dashboard.get_json() == {
        "error": "\u062f\u0633\u062a\u0631\u0633\u06cc \u063a\u06cc\u0631\u0645\u062c\u0627\u0632",
        "required_roles": ["admin"],
        "user_role": "expert",
    }

    unauthenticated_summary = client.get("/api/admin/reports/assignment-summary")
    assert unauthenticated_summary.status_code == 401
    assert unauthenticated_summary.get_json() == {"error": "Token is missing"}

    forbidden_summary = client.get(
        "/api/admin/reports/assignment-summary",
        headers=expert_headers,
    )
    assert forbidden_summary.status_code == 403
    assert forbidden_summary.get_json() == {
        "error": "دسترسی غیرمجاز",
        "required_roles": ["admin"],
        "user_role": "expert",
    }

    dashboard = client.get("/api/admin/dashboard", headers=admin_headers)
    assert dashboard.status_code == 200
    dashboard_payload = dashboard.get_json()
    assert set(dashboard_payload.keys()) == {
        "total_requests",
        "requests_per_transport_method",
        "requests_per_status",
        "last_7_days_count",
        "last_24h_count",
        "unassigned_count",
        "top_provinces",
    }
    assert dashboard_payload["total_requests"] == 3
    assert dashboard_payload["requests_per_transport_method"] == {
        "air": 1,
        "road": 1,
        "sea": 1,
    }
    assert dashboard_payload["requests_per_status"] == {"assigned": 1, "new": 1, "won": 1}
    assert dashboard_payload["last_7_days_count"] == 2
    assert dashboard_payload["last_24h_count"] == 1
    assert dashboard_payload["unassigned_count"] == 1
    assert dashboard_payload["top_provinces"] == [{"province": "Tehran", "count": 2}]

    summary = client.get("/api/admin/reports/assignment-summary", headers=admin_headers)
    assert summary.status_code == 200
    summary_payload = summary.get_json()
    assert set(summary_payload.keys()) == {"assignments_per_expert", "overall_stats", "generated_at"}
    assert len(summary_payload["assignments_per_expert"]) == 1
    expert_summary = summary_payload["assignments_per_expert"][0]
    assert expert_summary == {
        "expert_id": admin_panel_app["expert_id"],
        "expert_name": "Phase 5M Expert",
        "username": "phase5m_expert",
        "role": "expert",
        "total_assignments": 2,
        "won_count": 1,
        "lost_count": 0,
        "active_count": 1,
        "conversion_rate": 50.0,
        "avg_response_time_hours": None,
    }
    assert summary_payload["overall_stats"] == {
        "total_assignments": 2,
        "total_won": 1,
        "overall_conversion_rate": 50.0,
        "avg_response_time_hours": 1.0,
        "sla_violations": 1,
    }
    assert "T" in summary_payload["generated_at"]


@pytest.mark.skip(reason="ADR-043: reporting/export remains fail-closed pending a companion decision")
def test_admin_report_overview_auth_periods_sections_and_exclusions(admin_panel_app):
    """Admin report overview is admin-only, validates periods, and keeps the planned JSON shape."""
    client = admin_panel_app["app"].test_client()
    admin_headers = _auth_headers(admin_panel_app["admin_token"])
    expert_headers = _auth_headers(admin_panel_app["expert_token"])

    unauthenticated = client.get("/api/admin/reports/overview?period=weekly")
    assert unauthenticated.status_code == 401
    assert unauthenticated.get_json() == {"error": "Token is missing"}

    forbidden = client.get("/api/admin/reports/overview?period=weekly", headers=expert_headers)
    assert forbidden.status_code == 403
    assert forbidden.get_json() == {
        "error": "دسترسی غیرمجاز",
        "required_roles": ["admin"],
        "user_role": "expert",
    }

    invalid = client.get("/api/admin/reports/overview?period=invalid", headers=admin_headers)
    assert invalid.status_code == 400
    assert invalid.get_json() == {"error": "دوره گزارش نامعتبر است"}

    for period in ("weekly", "monthly", "yearly"):
        response = client.get(f"/api/admin/reports/overview?period={period}", headers=admin_headers)
        assert response.status_code == 200
        payload = response.get_json()
        assert set(payload.keys()) == {
            "metadata",
            "summary",
            "requests_by_status",
            "transport_distribution",
            "location_distribution",
            "expert_performance",
            "customer_concentration",
            "trends",
            "request_details",
        }
        assert payload["metadata"]["period"] == period
        assert payload["metadata"]["timezone_basis"] == "UTC"
        assert "T" in payload["metadata"]["generated_at"]
        _assert_report_exclusions(payload)


@pytest.mark.skip(reason="ADR-043: reporting/export remains fail-closed pending a companion decision")
def test_admin_report_overview_summary_labels_and_details(admin_panel_app):
    """Report overview returns status counts, labels, missing values, and detail rows without SLA/priority fields."""
    client = admin_panel_app["app"].test_client()
    admin_headers = _auth_headers(admin_panel_app["admin_token"])

    with admin_panel_app["app"].app_context():
        now = datetime.utcnow()
        customer = Customer(
            first_name="Linked",
            last_name="Customer",
            company_name="Acme Logistics",
            phone="09121111111",
        )
        db.session.add(customer)
        db.session.flush()

        in_progress_request = ShipmentRequest(
            tracking_code="AP-P12B-004",
            shipping_type="international",
            contact_phone="09121111111",
            customer_first_name="Linked",
            customer_last_name="Customer",
            customer_id=customer.id,
            international_transport_method="air",
            origin_country="Turkey",
            origin_city_international="Istanbul",
            dest_country="Iran",
            dest_city_international="Tehran",
            status_request_status="new",
            status="in_progress",
            assigned_to=admin_panel_app["expert_id"],
            created_at=now - timedelta(hours=3),
            cargo_description="International cargo",
            cargo_weight=12.5,
            cargo_volume=4.0,
            cargo_value=5000.0,
            has_unread_for_assignee=False,
        )
        lost_request = ShipmentRequest(
            tracking_code="AP-P12B-005",
            shipping_type="domestic",
            contact_phone="09122222222",
            status_request_status="new",
            status="lost",
            assigned_to=admin_panel_app["expert_id"],
            created_at=now - timedelta(hours=4),
        )
        missing_request = ShipmentRequest(
            tracking_code="AP-P12B-006",
            shipping_type="domestic",
            contact_phone="09123333333",
            status_request_status="new",
            status="new",
            assigned_to=None,
            created_at=now - timedelta(hours=5),
        )
        db.session.add_all([in_progress_request, lost_request, missing_request])
        db.session.commit()

    response = client.get("/api/admin/reports/overview?period=weekly", headers=admin_headers)
    assert response.status_code == 200
    payload = response.get_json()
    _assert_report_exclusions(payload)

    assert payload["summary"] == {
        "total_requests": 5,
        "new_requests": 1,
        "pending_review_requests": 1,
        "in_progress_requests": 1,
        "completed_requests": 1,
        "lost_or_rejected_requests": 1,
        "unassigned_requests": 1,
        "requests_last_24h": 4,
        "requests_last_7d": 5,
    }

    status_rows = {row["status"]: row for row in payload["requests_by_status"]}
    assert status_rows["assigned"]["label_fa"] == "در انتظار بررسی"
    assert status_rows["new"]["label_fa"] == "جدید"
    assert status_rows["in_progress"]["label_fa"] == "در حال بررسی"
    assert status_rows["lost"]["label_fa"] == "ردشده / از دست‌رفته"

    details_by_tracking = {row["tracking_number"]: row for row in payload["request_details"]}
    assert details_by_tracking["AP-P12B-006"]["expert_name"] == "بدون کارشناس"
    assert details_by_tracking["AP-P12B-006"]["customer_name"] == "ثبت نشده"
    assert details_by_tracking["AP-P12B-006"]["origin_province"] == "ثبت نشده"
    assert details_by_tracking["AP-P12B-006"]["transport_method"] == "unknown"
    assert details_by_tracking["AP-P12B-004"]["origin_province"] == "Turkey"
    assert details_by_tracking["AP-P12B-004"]["origin_city"] == "Istanbul"
    assert details_by_tracking["AP-P12B-004"]["destination_province"] == "Iran"
    assert details_by_tracking["AP-P12B-004"]["destination_city"] == "Tehran"

    expert_rows = {row["expert_name"]: row for row in payload["expert_performance"]}
    assert expert_rows["بدون کارشناس"]["assigned_count"] == 1
    assert expert_rows["Phase 5M Expert"]["assigned_count"] == 4

    customer_rows = {(row["customer_name"], row["phone"]): row for row in payload["customer_concentration"]}
    assert customer_rows[("Linked Customer", "09121111111")]["company"] == "Acme Logistics"

    assert len(payload["trends"]) == 7
    assert payload["request_details"][0]["created_at"] >= payload["request_details"][-1]["created_at"]


@pytest.mark.skip(reason="ADR-043: reporting/export remains fail-closed pending a companion decision")
def test_admin_report_overview_empty_report(admin_empty_report_app):
    """Report overview handles periods with no shipment requests safely."""
    client = admin_empty_report_app["app"].test_client()
    headers = _auth_headers(admin_empty_report_app["admin_token"])

    response = client.get("/api/admin/reports/overview?period=weekly", headers=headers)

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["summary"] == {
        "total_requests": 0,
        "new_requests": 0,
        "pending_review_requests": 0,
        "in_progress_requests": 0,
        "completed_requests": 0,
        "lost_or_rejected_requests": 0,
        "unassigned_requests": 0,
        "requests_last_24h": 0,
        "requests_last_7d": 0,
    }
    assert payload["requests_by_status"] == []
    assert payload["transport_distribution"] == []
    assert payload["location_distribution"] == []
    assert payload["expert_performance"] == []
    assert payload["customer_concentration"] == []
    assert payload["request_details"] == []
    assert len(payload["trends"]) == 7
    _assert_report_exclusions(payload)


@pytest.mark.skip(reason="ADR-043: reporting/export remains fail-closed pending a companion decision")
def test_admin_report_xlsx_export_auth_periods_headers_and_workbook(admin_panel_app):
    """Admin XLSX export is protected, validates periods, and returns a valid workbook."""
    client = admin_panel_app["app"].test_client()
    admin_headers = _auth_headers(admin_panel_app["admin_token"])
    expert_headers = _auth_headers(admin_panel_app["expert_token"])

    unauthenticated = client.get("/api/admin/reports/export.xlsx?period=weekly")
    assert unauthenticated.status_code == 401
    assert unauthenticated.get_json() == {"error": "Token is missing"}

    forbidden = client.get("/api/admin/reports/export.xlsx?period=weekly", headers=expert_headers)
    assert forbidden.status_code == 403
    assert forbidden.get_json() == {
        "error": "دسترسی غیرمجاز",
        "required_roles": ["admin"],
        "user_role": "expert",
    }

    invalid = client.get("/api/admin/reports/export.xlsx?period=invalid", headers=admin_headers)
    assert invalid.status_code == 400
    assert invalid.get_json() == {"error": "دوره گزارش نامعتبر است"}

    for period in ("weekly", "monthly", "yearly"):
        response = client.get(f"/api/admin/reports/export.xlsx?period={period}", headers=admin_headers)
        assert response.status_code == 200
        assert response.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers["Content-Disposition"]
        assert f"forwarder-report-{period}.xlsx" in response.headers["Content-Disposition"]
        assert load_workbook(BytesIO(response.data))


@pytest.mark.skip(reason="ADR-043: reporting/export remains fail-closed pending a companion decision")
def test_admin_report_xlsx_export_workbook_content(admin_panel_app):
    """XLSX export contains expected sheets, headers, labels, and summary metrics."""
    client = admin_panel_app["app"].test_client()
    admin_headers = _auth_headers(admin_panel_app["admin_token"])

    response = client.get("/api/admin/reports/export.xlsx?period=weekly", headers=admin_headers)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.data))
    expected_sheets = [
        "خلاصه مدیریتی",
        "وضعیت درخواست‌ها",
        "توزیع نوع حمل",
        "توزیع مبدا و مقصد",
        "عملکرد کارشناسان",
        "مشتریان و درخواست‌ها",
        "روند زمانی",
        "جزئیات درخواست‌ها",
    ]
    assert workbook.sheetnames == expected_sheets
    _assert_workbook_exclusions(workbook)

    summary = workbook["خلاصه مدیریتی"]
    summary_values = {
        summary.cell(row=row_index, column=1).value: summary.cell(row=row_index, column=2).value
        for row_index in range(1, summary.max_row + 1)
    }
    assert summary_values["کل درخواست‌ها"] == 2
    assert summary_values["در انتظار بررسی"] == 1
    assert summary_values["تکمیل‌شده"] == 1
    assert summary.freeze_panes == "A9"

    status_sheet = workbook["وضعیت درخواست‌ها"]
    assert [cell.value for cell in status_sheet[1]] == [
        "وضعیت",
        "برچسب فارسی وضعیت",
        "تعداد درخواست",
        "درصد از کل",
        "آخرین تاریخ ثبت در این وضعیت",
    ]
    status_rows = list(status_sheet.iter_rows(min_row=2, values_only=True))
    assert ("assigned", "در انتظار بررسی", 1, 50, status_rows[0][4]) in status_rows

    transport_sheet = workbook["توزیع نوع حمل"]
    assert [cell.value for cell in transport_sheet[1]] == [
        "نوع حمل",
        "برچسب فارسی نوع حمل",
        "تعداد درخواست",
        "درصد از کل",
        "تعداد تکمیل‌شده",
        "تعداد در حال بررسی",
    ]

    details_sheet = workbook["جزئیات درخواست‌ها"]
    assert "کارشناس" in [cell.value for cell in details_sheet[1]]
    assert "خوانده‌نشده برای کارشناس" in [cell.value for cell in details_sheet[1]]
    assert "sla_due_at" not in [cell.value for cell in details_sheet[1]]
    assert "priority" not in [cell.value for cell in details_sheet[1]]
    assert all(sheet.sheet_view.rightToLeft for sheet in workbook.worksheets)


@pytest.mark.skip(reason="ADR-043: reporting/export remains fail-closed pending a companion decision")
def test_admin_report_xlsx_export_empty_report(admin_empty_report_app):
    """XLSX export creates all sheets with headers when there is no report data."""
    client = admin_empty_report_app["app"].test_client()
    headers = _auth_headers(admin_empty_report_app["admin_token"])

    response = client.get("/api/admin/reports/export.xlsx?period=weekly", headers=headers)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.data))
    assert workbook.sheetnames == [
        "خلاصه مدیریتی",
        "وضعیت درخواست‌ها",
        "توزیع نوع حمل",
        "توزیع مبدا و مقصد",
        "عملکرد کارشناسان",
        "مشتریان و درخواست‌ها",
        "روند زمانی",
        "جزئیات درخواست‌ها",
    ]
    assert workbook["خلاصه مدیریتی"]["B9"].value == 0
    assert workbook["وضعیت درخواست‌ها"].max_row == 1
    assert workbook["جزئیات درخواست‌ها"].max_row == 1
    _assert_workbook_exclusions(workbook)
